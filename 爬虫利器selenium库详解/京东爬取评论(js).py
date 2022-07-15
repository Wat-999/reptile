#半成品

import requests
import json
import time
import re
import random
import pandas as pd
from faker import Faker
from multiprocessing import Pool
import os
from openpyxl import Workbook, load_workbook

from python爬虫.爬虫项目 import random_ip
ip = random_ip.dailiips()

#其中，json用于解析Json格式数据，faker用于生成随机浏览器请求头，multiprocessing为多进程相关模块，
#最终两行为笔者的的代理IP生成模块。免费的代理IP对京东来说作用不大，大家可以直接忽视该参数。

#2. 创建保存数据的文件

try:
    wb = load_workbook(filename='京东双十一华为手机评论信息.xlsx')
    sheet = wb.active
except:
    wb = Workbook()
    sheet = wb.active  # 激活工作簿
    sheet.title = '京东双十一华为手机评论信息'  # 设置工作簿名称
    # 设置表头
    sheet['A1'] = 'title'
    sheet['B1'] = 'user_id'
    sheet['C1'] = 'creationTime'
    sheet['D1'] = 'score'
    sheet['E1'] = 'content'

#3基本配置：主要是构造请求头和base网址，方便后续使用。
headers = {
    'User-Agent': Faker().chrome(),
    'Referer': 'https://item.jd.com/100011386554.html'
}

base_url = 'https://club.jd.com/comment/productPageComments.action?&productId=100011177233&score=1&sortType=5&page=1&pageSize=10&isShadowSku=0&fold=1'


#4 主体爬虫代码
def get_comment_id(start, end):
    '''
    读取京东商品的商品标题和sku（sku暂不知晓，记为商品评论唯一id）
    :param start: 开始的索引
    :param end: 结束的索引
    :return: 标题、skus列表
    '''
    df = pd.read_excel(r'C:\Users\HTHDS\Desktop\python\python爬虫\爬虫项目\京东双十一华为手机信息.xlsx')
    names = df['name'][start:end]
    skus = df['sku'][start:end]
    return names, skus

def get_maxpage(sku):
    '''
    获取每类评论（好评、中评和差评）的最大页数
    :param sku: 商品评论唯一id
    :return:
    '''
    global base_url
    maxpagelist = []
    for score in [1, 2, 3]:
        url = base_url.replace('score=1', f'score={score}').replace('productId=100011177233', f'productId={sku}')
        print(url)
        response = requests.get(url, proxies=random_ip.dailiips(), headers=headers).text
        html = response
        json_data = json.loads(html)
        maxpagelist.append(int(json_data['maxPage']))
    return maxpagelist

def get_url(score, page, sku, title):
    '''
    获取每页评论的评论数据
    :param score: 评论类别
    :param page: 页码
    :param sku: 商品唯一id
    :param title: 标题
    :return: 获取到的评论数据
    '''
    global base_url
    print('正在爬取第{}页,id={}'.format(page, os.getpid()))
    url = base_url.replace('page=1', f'page={page}').replace('score=1', f'score={score}').replace('productId=100011177233', f'productId={sku}')
    data_list_list = get_comment(url, headers=headers, title=title)  # 调用采集方法
    return data_list_list

def get_comment(url=None, headers=None, title=''):
    '''
    采集评论数据
    :param url: 评论链接
    :param headers: 请求头
    :param title: 标题
    :return: 获取到的评论数据
    '''
    data_list_list = []
    response = requests.get(url, proxies=random_ip.dailiips(), headers=headers).text
    html = response
    json_data = json.loads(html)
    comment = json_data['comments']
    for data in comment:
        user_id = data['id']   # 评论用户id
        creationTime = data['creationTime']   # 评论时间
        content = data['content'].strip()   # 评论文本
        content = re.sub('\n','',content)
        score = data['score']  # 评分
        data_item = [title, user_id, creationTime, score, content]
        data_list_list.append(data_item)
        print(data_item)
    return data_list_list

def write_comment(data_list_list=None):
   for data in data_list_list:
       sheet.append(data)



#开启爬虫程序
def main():
    start = time.time()
    # 创建10个进程
    pool = Pool(10)
    names, skus = get_comment_id(13, 14)
    for name, sku in zip(names, skus):
        print(name)
        sp = get_maxpage(sku)
        for s, v in enumerate(sp):
            for p in range(int(v)):
                pool.apply_async(get_url, args=(s+1, p, sku, name), callback=write_comment)
                time.sleep(random.randint(5, 8))
        time.sleep(random.randint(60, 80))
    pool.close()  # 无任务就关闭所有子进程
    pool.join()  # 子进程未结束阻止主进程执行下面程序
    wb.save(filename='京东双十一华为手机评论信息.xlsx')
    end = time.time()
    print('此次爬完数据用时：', (end-start))


if __name__ == '__main__':
    main()









